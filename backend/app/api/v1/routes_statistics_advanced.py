from __future__ import annotations

import io
import json
from datetime import UTC, datetime
from typing import Any, Literal
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import HTMLResponse, StreamingResponse
from openpyxl import Workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.dependencies import get_current_user, require_roles
from app.db.session import get_db_session
from app.models.auth import Membership, OrganizationRole, User
from app.models.statistics import (
    AnalysisStatus,
    StatisticalAnalysis,
    StatisticalChart,
    StatisticalDataset,
    StatisticalMethodComparison,
    StatisticalReport,
    StatisticalReportRevision,
    StatisticalReviewComment,
    StatisticalSampleSizePlan,
    StatisticalSensitivityRun,
    StatisticalStudy,
)
from app.schemas.statistics import (
    AnalysisRead,
    AnalysisVersionCreate,
    MethodComparisonRead,
    PValueAdjustmentRead,
    PValueAdjustmentRequest,
    MethodComparisonRequest,
    ReportRevisionCreate,
    ReportRevisionRead,
    ReviewCommentCreate,
    ReviewCommentRead,
    ReviewStatusUpdate,
    SampleSizePlanCreate,
    SampleSizePlanRead,
    ScriptExportRead,
    SensitivityRead,
    SensitivityRequest,
)
from app.services.statistical_exports import (
    render_chart_bytes,
    report_docx_bytes,
    report_pdf_bytes,
)
from app.services.statistics_advanced import (
    calculate_sample_size,
    compare_methods,
    configuration_checksum,
    generate_reproduction_script,
    result_signature,
    run_sensitivity_scenario,
)
from app.services.statistics_engine import adjust_p_values, execute

router = APIRouter(prefix="/statistics", tags=["Laboratório Estatístico Avançado"])
ROLES = (OrganizationRole.OWNER, OrganizationRole.ADMIN, OrganizationRole.TEACHER)


def _organization_id(membership: Membership) -> UUID:
    return membership.organization_id


async def _analysis_or_404(
    analysis_id: UUID, session: AsyncSession, organization_id: UUID
) -> StatisticalAnalysis:
    analysis = await session.scalar(
        select(StatisticalAnalysis).where(
            StatisticalAnalysis.id == analysis_id,
            StatisticalAnalysis.organization_id == organization_id,
        )
    )
    if analysis is None:
        raise HTTPException(404, "Análise não encontrada")
    return analysis


async def _report_or_404(
    report_id: UUID, session: AsyncSession, organization_id: UUID
) -> StatisticalReport:
    report = await session.scalar(
        select(StatisticalReport).where(
            StatisticalReport.id == report_id,
            StatisticalReport.organization_id == organization_id,
        )
    )
    if report is None:
        raise HTTPException(404, "Relatório não encontrado")
    return report


@router.post(
    "/analyses/{analysis_id}/versions",
    response_model=AnalysisRead,
    status_code=201,
)
async def create_analysis_version(
    analysis_id: UUID,
    data: AnalysisVersionCreate,
    session: AsyncSession = Depends(get_db_session),
    membership: Membership = Depends(require_roles(*ROLES)),
    user: User = Depends(get_current_user),
) -> StatisticalAnalysis:
    organization_id = _organization_id(membership)
    base = await _analysis_or_404(analysis_id, session, organization_id)
    dataset = await session.scalar(
        select(StatisticalDataset).where(
            StatisticalDataset.id == base.dataset_id,
            StatisticalDataset.organization_id == organization_id,
        )
    )
    study = await session.get(StatisticalStudy, base.study_id)
    if dataset is None or study is None:
        raise HTTPException(404, "Dataset ou estudo não encontrado")
    parameters = data.parameters if data.parameters is not None else base.parameters
    latest_version = await session.scalar(
        select(func.max(StatisticalAnalysis.version_number)).where(
            (StatisticalAnalysis.id == base.id)
            | (StatisticalAnalysis.parent_analysis_id == base.id)
        )
    )
    version = StatisticalAnalysis(
        study_id=base.study_id,
        dataset_id=base.dataset_id,
        organization_id=organization_id,
        title=data.title or f"{base.title} — versão {(latest_version or 1) + 1}",
        analysis_type=base.analysis_type,
        parameters=parameters,
        created_by_user_id=user.id,
        parent_analysis_id=base.id,
        version_number=(latest_version or 1) + 1,
        configuration_checksum=configuration_checksum(
            dataset.dataset_checksum, base.analysis_type, parameters
        ),
        status=AnalysisStatus.PENDING,
        software_versions={
            "engine": "educode-statistics-1.1",
            "python": "3.12",
            "change_summary": data.change_summary,
        },
    )
    session.add(version)
    await session.flush()
    try:
        result = execute(
            dataset.rows_snapshot,
            version.analysis_type,
            version.parameters,
            study.significance_level,
        )
        for key, value in result.items():
            setattr(version, key, value)
        version.result_signature = result_signature(result)
        version.status = AnalysisStatus.COMPLETED
        version.executed_at = datetime.now(UTC)
    except (ValueError, RuntimeError) as exc:
        version.status = AnalysisStatus.FAILED
        version.limitations = [str(exc)]
        await session.commit()
        raise HTTPException(422, str(exc)) from exc
    await session.commit()
    await session.refresh(version)
    return version


@router.post(
    "/analyses/{analysis_id}/sensitivity",
    response_model=list[SensitivityRead],
    status_code=201,
)
async def run_sensitivity(
    analysis_id: UUID,
    data: SensitivityRequest,
    session: AsyncSession = Depends(get_db_session),
    membership: Membership = Depends(require_roles(*ROLES)),
    user: User = Depends(get_current_user),
) -> list[StatisticalSensitivityRun]:
    organization_id = _organization_id(membership)
    analysis = await _analysis_or_404(analysis_id, session, organization_id)
    dataset = await session.get(StatisticalDataset, analysis.dataset_id)
    study = await session.get(StatisticalStudy, analysis.study_id)
    if dataset is None or study is None:
        raise HTTPException(404, "Dataset ou estudo não encontrado")
    base_significant = analysis.test_results.get("significant")
    created: list[StatisticalSensitivityRun] = []
    for scenario_key in data.scenario_keys:
        try:
            method, transformed_rows, result, scenario_parameters = run_sensitivity_scenario(
                dataset.rows_snapshot,
                analysis.analysis_type,
                analysis.parameters,
                study.significance_level,
                scenario_key,
                data.alternative_analysis_type,
            )
        except (ValueError, RuntimeError) as exc:
            raise HTTPException(422, str(exc)) from exc
        scenario_significant = result.get("test_results", {}).get("significant")
        run = StatisticalSensitivityRun(
            base_analysis_id=analysis.id,
            dataset_id=dataset.id,
            organization_id=organization_id,
            title=f"{analysis.title} — {scenario_key}",
            scenario_key=scenario_key,
            scenario_parameters={
                **scenario_parameters,
                "dataset_checksum": dataset.dataset_checksum,
            },
            analysis_type=method,
            result={**result, "analyzed_rows": len(transformed_rows)},
            conclusion_changed=(
                base_significant is not None
                and scenario_significant is not None
                and bool(base_significant) != bool(scenario_significant)
            ),
            created_by_user_id=user.id,
        )
        session.add(run)
        created.append(run)
    await session.commit()
    for run in created:
        await session.refresh(run)
    return created


@router.get(
    "/analyses/{analysis_id}/sensitivity",
    response_model=list[SensitivityRead],
)
async def list_sensitivity_runs(
    analysis_id: UUID,
    session: AsyncSession = Depends(get_db_session),
    membership: Membership = Depends(require_roles(*ROLES)),
) -> list[StatisticalSensitivityRun]:
    organization_id = _organization_id(membership)
    await _analysis_or_404(analysis_id, session, organization_id)
    return list(
        (
            await session.scalars(
                select(StatisticalSensitivityRun)
                .where(
                    StatisticalSensitivityRun.base_analysis_id == analysis_id,
                    StatisticalSensitivityRun.organization_id == organization_id,
                )
                .order_by(StatisticalSensitivityRun.created_at.desc())
            )
        ).all()
    )


@router.post(
    "/analyses/{analysis_id}/method-comparisons",
    response_model=MethodComparisonRead,
    status_code=201,
)
async def create_method_comparison(
    analysis_id: UUID,
    data: MethodComparisonRequest,
    session: AsyncSession = Depends(get_db_session),
    membership: Membership = Depends(require_roles(*ROLES)),
    user: User = Depends(get_current_user),
) -> StatisticalMethodComparison:
    organization_id = _organization_id(membership)
    analysis = await _analysis_or_404(analysis_id, session, organization_id)
    dataset = await session.get(StatisticalDataset, analysis.dataset_id)
    study = await session.get(StatisticalStudy, analysis.study_id)
    if dataset is None or study is None:
        raise HTTPException(404, "Dataset ou estudo não encontrado")
    try:
        comparison_result = compare_methods(
            dataset.rows_snapshot,
            data.methods,
            analysis.parameters,
            study.significance_level,
        )
    except (ValueError, RuntimeError) as exc:
        raise HTTPException(422, str(exc)) from exc
    comparison = StatisticalMethodComparison(
        base_analysis_id=analysis.id,
        dataset_id=dataset.id,
        organization_id=organization_id,
        methods=data.methods,
        results=comparison_result["results"],
        recommendation=comparison_result["recommendation"],
        conclusions_consistent=comparison_result["conclusions_consistent"],
        created_by_user_id=user.id,
    )
    session.add(comparison)
    await session.commit()
    await session.refresh(comparison)
    return comparison


@router.get(
    "/analyses/{analysis_id}/method-comparisons",
    response_model=list[MethodComparisonRead],
)
async def list_method_comparisons(
    analysis_id: UUID,
    session: AsyncSession = Depends(get_db_session),
    membership: Membership = Depends(require_roles(*ROLES)),
) -> list[StatisticalMethodComparison]:
    organization_id = _organization_id(membership)
    await _analysis_or_404(analysis_id, session, organization_id)
    return list(
        (
            await session.scalars(
                select(StatisticalMethodComparison)
                .where(
                    StatisticalMethodComparison.base_analysis_id == analysis_id,
                    StatisticalMethodComparison.organization_id == organization_id,
                )
                .order_by(StatisticalMethodComparison.created_at.desc())
            )
        ).all()
    )


@router.get(
    "/analyses/{analysis_id}/scripts/{language}",
    response_model=ScriptExportRead,
)
async def export_reproduction_script(
    analysis_id: UUID,
    language: Literal["python", "r"],
    session: AsyncSession = Depends(get_db_session),
    membership: Membership = Depends(require_roles(*ROLES)),
) -> dict[str, Any]:
    organization_id = _organization_id(membership)
    analysis = await _analysis_or_404(analysis_id, session, organization_id)
    dataset = await session.get(StatisticalDataset, analysis.dataset_id)
    if dataset is None:
        raise HTTPException(404, "Dataset não encontrado")
    content = generate_reproduction_script(
        language,
        analysis.analysis_type,
        analysis.parameters,
        dataset.dataset_checksum,
    )
    extension = "py" if language == "python" else "R"
    return {
        "language": language,
        "filename": f"analysis-{analysis.id}.{extension}",
        "dataset_checksum": dataset.dataset_checksum,
        "content": content,
    }


@router.get("/analyses/{analysis_id}/scripts/{language}/download")
async def download_reproduction_script(
    analysis_id: UUID,
    language: Literal["python", "r"],
    session: AsyncSession = Depends(get_db_session),
    membership: Membership = Depends(require_roles(*ROLES)),
) -> StreamingResponse:
    payload = await export_reproduction_script(analysis_id, language, session, membership)
    return StreamingResponse(
        iter([payload["content"]]),
        media_type="text/x-python" if language == "python" else "text/plain",
        headers={"Content-Disposition": f'attachment; filename="{payload["filename"]}"'},
    )


@router.post("/p-values/adjust", response_model=PValueAdjustmentRead)
async def adjust_multiple_p_values(
    data: PValueAdjustmentRequest,
    membership: Membership = Depends(require_roles(*ROLES)),
) -> dict[str, Any]:
    _ = membership
    return {
        "method": data.method,
        "original_p_values": data.p_values,
        "adjusted_p_values": adjust_p_values(data.p_values, data.method),
    }


@router.get("/datasets/{dataset_id}/xlsx")
async def dataset_xlsx(
    dataset_id: UUID,
    session: AsyncSession = Depends(get_db_session),
    membership: Membership = Depends(require_roles(*ROLES)),
) -> StreamingResponse:
    dataset = await session.scalar(
        select(StatisticalDataset).where(
            StatisticalDataset.id == dataset_id,
            StatisticalDataset.organization_id == _organization_id(membership),
        )
    )
    if dataset is None:
        raise HTTPException(404, "Dataset não encontrado")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dados anonimizados"
    keys = sorted({key for row in dataset.rows_snapshot for key in row})
    sheet.append(keys)
    for row in dataset.rows_snapshot:
        values = []
        for key in keys:
            value = row.get(key)
            if isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False, default=str)
            values.append(value)
        sheet.append(values)
    dictionary = workbook.create_sheet("Dicionário")
    dictionary.append(["Variável", "Tipo", "Descrição"])
    for item in dataset.variable_dictionary:
        dictionary.append([item.get("name"), item.get("type"), item.get("description")])
    metadata = workbook.create_sheet("Metadados")
    metadata.append(["dataset_checksum", dataset.dataset_checksum])
    metadata.append(["attempt_policy", dataset.attempt_policy])
    metadata.append(["snapshot_at", dataset.snapshot_at.isoformat()])
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="dataset-{dataset.id}.xlsx"'
        },
    )


@router.post(
    "/sample-size-plans",
    response_model=SampleSizePlanRead,
    status_code=201,
)
async def create_sample_size_plan(
    data: SampleSizePlanCreate,
    session: AsyncSession = Depends(get_db_session),
    membership: Membership = Depends(require_roles(*ROLES)),
    user: User = Depends(get_current_user),
) -> StatisticalSampleSizePlan:
    if data.study_id is not None:
        study = await session.scalar(
            select(StatisticalStudy).where(
                StatisticalStudy.id == data.study_id,
                StatisticalStudy.organization_id == _organization_id(membership),
            )
        )
        if study is None:
            raise HTTPException(404, "Estudo não encontrado")
    try:
        result = calculate_sample_size(
            data.design,
            data.significance_level,
            data.power,
            data.expected_effect_size,
            data.group_ratio,
            data.parameters,
        )
    except ValueError as exc:
        raise HTTPException(422, str(exc)) from exc
    plan = StatisticalSampleSizePlan(
        organization_id=_organization_id(membership),
        created_by_user_id=user.id,
        result=result,
        **data.model_dump(),
    )
    session.add(plan)
    await session.commit()
    await session.refresh(plan)
    return plan


@router.get("/sample-size-plans", response_model=list[SampleSizePlanRead])
async def list_sample_size_plans(
    session: AsyncSession = Depends(get_db_session),
    membership: Membership = Depends(require_roles(*ROLES)),
) -> list[StatisticalSampleSizePlan]:
    return list(
        (
            await session.scalars(
                select(StatisticalSampleSizePlan)
                .where(
                    StatisticalSampleSizePlan.organization_id
                    == _organization_id(membership)
                )
                .order_by(StatisticalSampleSizePlan.created_at.desc())
            )
        ).all()
    )


async def _validate_review_entity(
    entity_type: str,
    entity_id: UUID,
    session: AsyncSession,
    organization_id: UUID,
) -> None:
    if entity_type == "analysis":
        await _analysis_or_404(entity_id, session, organization_id)
    elif entity_type == "report":
        await _report_or_404(entity_id, session, organization_id)
    else:
        raise HTTPException(422, "Tipo de entidade de revisão inválido")


@router.post(
    "/review-comments",
    response_model=ReviewCommentRead,
    status_code=201,
)
async def create_review_comment(
    data: ReviewCommentCreate,
    session: AsyncSession = Depends(get_db_session),
    membership: Membership = Depends(require_roles(*ROLES)),
    user: User = Depends(get_current_user),
) -> StatisticalReviewComment:
    organization_id = _organization_id(membership)
    await _validate_review_entity(
        data.entity_type, data.entity_id, session, organization_id
    )
    comment = StatisticalReviewComment(
        organization_id=organization_id,
        created_by_user_id=user.id,
        **data.model_dump(),
    )
    session.add(comment)
    await session.commit()
    await session.refresh(comment)
    return comment


@router.get("/review-comments", response_model=list[ReviewCommentRead])
async def list_review_comments(
    entity_type: Literal["analysis", "report"] = Query(...),
    entity_id: UUID = Query(...),
    session: AsyncSession = Depends(get_db_session),
    membership: Membership = Depends(require_roles(*ROLES)),
) -> list[StatisticalReviewComment]:
    organization_id = _organization_id(membership)
    await _validate_review_entity(entity_type, entity_id, session, organization_id)
    return list(
        (
            await session.scalars(
                select(StatisticalReviewComment)
                .where(
                    StatisticalReviewComment.organization_id == organization_id,
                    StatisticalReviewComment.entity_type == entity_type,
                    StatisticalReviewComment.entity_id == entity_id,
                )
                .order_by(StatisticalReviewComment.created_at)
            )
        ).all()
    )


@router.patch(
    "/review-comments/{comment_id}/resolve",
    response_model=ReviewCommentRead,
)
async def resolve_review_comment(
    comment_id: UUID,
    session: AsyncSession = Depends(get_db_session),
    membership: Membership = Depends(require_roles(*ROLES)),
    user: User = Depends(get_current_user),
) -> StatisticalReviewComment:
    comment = await session.scalar(
        select(StatisticalReviewComment).where(
            StatisticalReviewComment.id == comment_id,
            StatisticalReviewComment.organization_id == _organization_id(membership),
        )
    )
    if comment is None:
        raise HTTPException(404, "Comentário não encontrado")
    comment.status = "resolved"
    comment.resolved_by_user_id = user.id
    comment.resolved_at = datetime.now(UTC)
    await session.commit()
    await session.refresh(comment)
    return comment


@router.patch("/analyses/{analysis_id}/review-status", response_model=AnalysisRead)
async def update_analysis_review_status(
    analysis_id: UUID,
    data: ReviewStatusUpdate,
    session: AsyncSession = Depends(get_db_session),
    membership: Membership = Depends(require_roles(*ROLES)),
) -> StatisticalAnalysis:
    analysis = await _analysis_or_404(
        analysis_id, session, _organization_id(membership)
    )
    analysis.review_status = data.status
    if data.status == "in_review":
        analysis.status = AnalysisStatus.IN_REVIEW
    elif data.status == "approved":
        analysis.status = AnalysisStatus.APPROVED
    await session.commit()
    await session.refresh(analysis)
    return analysis


@router.get("/analyses/{analysis_id}/reports")
async def list_analysis_reports(
    analysis_id: UUID,
    session: AsyncSession = Depends(get_db_session),
    membership: Membership = Depends(require_roles(*ROLES)),
) -> list[dict[str, Any]]:
    organization_id = _organization_id(membership)
    await _analysis_or_404(analysis_id, session, organization_id)
    reports = list(
        (
            await session.scalars(
                select(StatisticalReport)
                .where(
                    StatisticalReport.analysis_id == analysis_id,
                    StatisticalReport.organization_id == organization_id,
                )
                .order_by(StatisticalReport.created_at.desc())
            )
        ).all()
    )
    return [
        {
            "id": report.id,
            "title": report.title,
            "report_type": report.report_type,
            "version_number": report.version_number,
            "review_status": report.review_status,
            "created_at": report.created_at,
        }
        for report in reports
    ]


@router.post(
    "/reports/{report_id}/revisions",
    response_model=ReportRevisionRead,
    status_code=201,
)
async def create_report_revision(
    report_id: UUID,
    data: ReportRevisionCreate,
    session: AsyncSession = Depends(get_db_session),
    membership: Membership = Depends(require_roles(*ROLES)),
    user: User = Depends(get_current_user),
) -> StatisticalReportRevision:
    report = await _report_or_404(
        report_id, session, _organization_id(membership)
    )
    next_version = report.version_number + 1
    if data.title:
        report.title = data.title
    report.content_html = data.content_html or report.content_html
    report.sections = data.sections if data.sections is not None else report.sections
    report.version_number = next_version
    report.review_status = "draft"
    revision = StatisticalReportRevision(
        report_id=report.id,
        organization_id=report.organization_id,
        version_number=next_version,
        content_html=report.content_html,
        sections=report.sections,
        change_summary=data.change_summary,
        created_by_user_id=user.id,
    )
    session.add(revision)
    await session.commit()
    await session.refresh(revision)
    return revision


@router.get(
    "/reports/{report_id}/revisions",
    response_model=list[ReportRevisionRead],
)
async def list_report_revisions(
    report_id: UUID,
    session: AsyncSession = Depends(get_db_session),
    membership: Membership = Depends(require_roles(*ROLES)),
) -> list[StatisticalReportRevision]:
    organization_id = _organization_id(membership)
    await _report_or_404(report_id, session, organization_id)
    return list(
        (
            await session.scalars(
                select(StatisticalReportRevision)
                .where(
                    StatisticalReportRevision.report_id == report_id,
                    StatisticalReportRevision.organization_id == organization_id,
                )
                .order_by(StatisticalReportRevision.version_number.desc())
            )
        ).all()
    )


@router.patch("/reports/{report_id}/review-status")
async def update_report_review_status(
    report_id: UUID,
    data: ReviewStatusUpdate,
    session: AsyncSession = Depends(get_db_session),
    membership: Membership = Depends(require_roles(*ROLES)),
) -> dict[str, Any]:
    report = await _report_or_404(
        report_id, session, _organization_id(membership)
    )
    report.review_status = data.status
    await session.commit()
    return {
        "id": report.id,
        "review_status": report.review_status,
        "version_number": report.version_number,
    }


@router.get("/charts/{chart_id}/export/{output_format}")
async def export_chart(
    chart_id: UUID,
    output_format: Literal["png", "svg", "pdf"],
    session: AsyncSession = Depends(get_db_session),
    membership: Membership = Depends(require_roles(*ROLES)),
) -> StreamingResponse:
    chart = await session.scalar(
        select(StatisticalChart).where(
            StatisticalChart.id == chart_id,
            StatisticalChart.organization_id == _organization_id(membership),
        )
    )
    if chart is None:
        raise HTTPException(404, "Gráfico não encontrado")
    content = render_chart_bytes(
        chart.chart_type,
        chart.data_snapshot,
        chart.configuration,
        chart.title,
        chart.description,
        output_format,
    )
    media_types = {
        "png": "image/png",
        "svg": "image/svg+xml",
        "pdf": "application/pdf",
    }
    return StreamingResponse(
        io.BytesIO(content),
        media_type=media_types[output_format],
        headers={
            "Content-Disposition": (
                f'attachment; filename="chart-{chart.id}.{output_format}"'
            )
        },
    )


@router.get("/reports/{report_id}/download/{output_format}")
async def download_report(
    report_id: UUID,
    output_format: Literal["html", "pdf", "docx"],
    session: AsyncSession = Depends(get_db_session),
    membership: Membership = Depends(require_roles(*ROLES)),
) -> Any:
    organization_id = _organization_id(membership)
    report = await _report_or_404(report_id, session, organization_id)
    analysis = await _analysis_or_404(report.analysis_id, session, organization_id)
    dataset = await session.get(StatisticalDataset, analysis.dataset_id)
    charts = list(
        (
            await session.scalars(
                select(StatisticalChart)
                .where(
                    StatisticalChart.analysis_id == analysis.id,
                    StatisticalChart.organization_id == organization_id,
                    StatisticalChart.include_in_report.is_(True),
                )
                .order_by(StatisticalChart.display_order)
            )
        ).all()
    )
    metadata = {
        "dataset_checksum": dataset.dataset_checksum if dataset else "",
        "analysis_signature": analysis.result_signature,
        "configuration_checksum": analysis.configuration_checksum,
        "analysis_version": analysis.version_number,
        "report_version": report.version_number,
    }
    if output_format == "html":
        return HTMLResponse(report.content_html)
    if output_format == "pdf":
        content = report_pdf_bytes(
            report.title, report.content_html, charts, metadata
        )
        media_type = "application/pdf"
    else:
        content = report_docx_bytes(
            report.title, report.content_html, charts, metadata
        )
        media_type = (
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        )
    return StreamingResponse(
        io.BytesIO(content),
        media_type=media_type,
        headers={
            "Content-Disposition": (
                f'attachment; filename="report-{report.id}.{output_format}"'
            )
        },
    )
